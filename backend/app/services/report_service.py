"""
services/report_service.py – PDF and Excel report generation.
"""
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path

from loguru import logger
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.repositories.transaction_repository import TransactionRepository
from app.models.transaction import TransactionType

settings = get_settings()


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.tx_repo = TransactionRepository(db)

    async def generate_monthly_pdf(
        self,
        family_id: uuid.UUID,
        year: int,
        month: int,
    ) -> bytes:
        """Generate a monthly summary PDF and return raw bytes."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors

        date_from = datetime(year, month, 1)
        date_to = datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1)

        income = await self.tx_repo.sum_by_type(family_id, TransactionType.INCOME, date_from, date_to)
        expenses = await self.tx_repo.sum_by_type(family_id, TransactionType.EXPENSE, date_from, date_to)
        net = income - expenses

        cats = await self.tx_repo.sum_by_category(family_id, date_from, date_to)
        transactions = await self.tx_repo.get_by_date_range(family_id, date_from, date_to)

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        story = []

        # Title
        story.append(Paragraph(f"Family Office – Monthly Report", styles["Title"]))
        story.append(Paragraph(f"{datetime(year, month, 1).strftime('%B %Y')}", styles["Heading2"]))
        story.append(Spacer(1, 10))

        # Summary table
        summary_data = [
            ["Metric", "Amount"],
            ["Total Income", f"${income:,.2f}"],
            ["Total Expenses", f"${expenses:,.2f}"],
            ["Net Savings", f"${net:,.2f}"],
            ["Savings Rate", f"{float(net/income*100) if income else 0:.1f}%"],
        ]
        t = Table(summary_data, colWidths=[80*mm, 60*mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#3d5c41")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f0ede8")]),
            ("ALIGN", (1,0), (1,-1), "RIGHT"),
        ]))
        story.append(t)
        story.append(Spacer(1, 10))

        # Transactions
        story.append(Paragraph("Transaction Detail", styles["Heading3"]))
        tx_data = [["Date", "Description", "Category", "Amount"]]
        for tx in transactions[:50]:
            tx_data.append([
                tx.transaction_date.strftime("%b %d"),
                tx.description[:35],
                "—",
                f"{'+'if tx.type==TransactionType.INCOME else '-'}${abs(float(tx.amount)):,.2f}",
            ])
        tx_table = Table(tx_data, colWidths=[25*mm, 80*mm, 40*mm, 30*mm])
        tx_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2a2a24")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f0ede8")]),
        ]))
        story.append(tx_table)

        doc.build(story)
        return buf.getvalue()

    async def generate_excel_export(
        self,
        family_id: uuid.UUID,
        date_from: datetime,
        date_to: datetime,
    ) -> bytes:
        """Export transactions as Excel workbook."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        transactions = await self.tx_repo.get_by_date_range(family_id, date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        header_fill = PatternFill(start_color="3d5c41", end_color="3d5c41", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        headers = ["Date", "Description", "Type", "Amount", "Currency", "Category", "Member", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, tx in enumerate(transactions, 2):
            ws.cell(row=row_idx, column=1, value=tx.transaction_date.strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=2, value=tx.description)
            ws.cell(row=row_idx, column=3, value=tx.type.value)
            ws.cell(row=row_idx, column=4, value=float(tx.amount))
            ws.cell(row=row_idx, column=5, value=tx.currency)
            ws.cell(row=row_idx, column=6, value=str(tx.category_id or ""))
            ws.cell(row=row_idx, column=7, value=str(tx.user_id or ""))
            ws.cell(row=row_idx, column=8, value=tx.notes or "")

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
